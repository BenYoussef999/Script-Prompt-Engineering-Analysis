import os
import json
import random
import datetime
from openai import OpenAI
from openpyxl import Workbook, load_workbook

# CONFIG 
EXCEL_FILE = "prompt.xlsx"
QUESTIONS_FILE = "questions.json"
MODEL = "openai/gpt-oss-120b:free"
BATCH_SIZE = 15

#API 
api_key = "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"

client = OpenAI(
    api_key=api_key,
    base_url="https://openrouter.ai/api/v1",
)

if not api_key:
    raise ValueError("Missing OPENROUTER_API_KEY")

client = OpenAI(
    api_key=api_key,
    base_url="https://openrouter.ai/api/v1",
)

# QUESTION GENERATORS

def gen_math():
    a, b = random.randint(1, 200), random.randint(1, 200)
    op = random.choice(["+", "-", "*"])
    if op == "+":
        return f"What is {a} + {b}?", str(a + b)
    elif op == "-":
        return f"What is {a} - {b}?", str(a - b)
    else:
        return f"What is {a} × {b}?", str(a * b)

def gen_science():
    return random.choice([
        ("What gas do humans breathe in?", "Oxygen"),
        ("What is H2O?", "Water"),
        ("What planet is closest to the Sun?", "Mercury"),
        ("What do plants absorb?", "Carbon dioxide"),
    ])

def gen_history():
    return random.choice([
        ("In what year did WW2 end?", "1945"),
        ("Who was the first US president?", "George Washington"),
        ("In what year did WW1 end?", "1918"),
    ])

def gen_geography():
    return random.choice([
        ("What is the capital of France?", "Paris"),
        ("What is the largest continent?", "Asia"),
        ("How many continents are there?", "7"),
    ])

def gen_physics():
    return random.choice([
        ("What is gravity?", "Force pulling objects down"),
        ("What is F = ma?", "Newton's Second Law"),
        ("Speed of light?", "299792 km/s"),
    ])

GENERATORS = [
    ("Math", gen_math),
    ("Science", gen_science),
    ("History", gen_history),
    ("Geography", gen_geography),
    ("Physics", gen_physics),
]

#  GENERATE QUESTIONS FILE 

def generate_questions_file():
    if os.path.exists(QUESTIONS_FILE):
        return

    questions = []
    seen = set()

    while len(questions) < 100:
        cat, func = random.choice(GENERATORS)
        q, a = func()

        if q in seen:
            continue

        seen.add(q)

        questions.append({
            "question": q,
            "answer": a,
            "category": cat
        })

    with open(QUESTIONS_FILE, "w") as f:
        json.dump(questions, f, indent=4)

    print(" questions.json created with 100 questions")

# LOAD QUESTIONS

def load_questions():
    with open(QUESTIONS_FILE, "r") as f:
        return json.load(f)

# PROMPTS

def zero_shot(q):
    return q

def cot(q):
    return f"Solve step by step.\n\nQuestion: {q}\n\nFinal answer:"

def few_shot(q, category):
    examples = {
        "Math": "Q: 2+2?\nA: 4\n\nQ: 3+3?\nA: 6\n\n",
        "Science": "Q: H2O?\nA: Water\n\nQ: Gas humans breathe?\nA: Oxygen\n\n",
        "History": "Q: WW2 ended?\nA: 1945\n\n",
        "Geography": "Q: Capital of France?\nA: Paris\n\n",
        "Physics": "Q: Gravity?\nA: Force pulling objects down\n\n",
    }
    return examples.get(category, "") + f"Q: {q}\nA:"

TECHNIQUES = {
    "Zero-Shot": zero_shot,
    "Few-Shot": few_shot,
    "Chain-of-Thought": cot,
}

# LLM CALL

def ask_llm(prompt):
    response = client.chat.completions.create(
        model=MODEL,
        messages=[{"role": "user", "content": prompt}],
        max_tokens=500,
        temperature=0,
    )
    return response.choices[0].message.content.strip()

#  EXCEL SETUP

HEADERS = [
    "Index", "Timestamp", "Category", "Question", "Answer",
    "Technique", "Prompt", "Response"
]

def setup_excel():
    if not os.path.exists(EXCEL_FILE):
        wb = Workbook()
        ws = wb.active
        ws.title = "Results"
        ws.append(HEADERS)
        wb.save(EXCEL_FILE)

    wb = load_workbook(EXCEL_FILE)
    return wb, wb.active

# PROGRESS TRACKING 

def get_completed_questions(ws):
    return (ws.max_row - 1) // 3

# MAIN

def main():
    generate_questions_file()
    questions = load_questions()

    wb, ws = setup_excel()

    done = get_completed_questions(ws)
    start = done
    end = min(done + BATCH_SIZE, len(questions))

    print(f"\n Running questions {start+1} → {end}")

    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    for i in range(start, end):
        item = questions[i]
        q = item["question"]
        a = item["answer"]
        cat = item["category"]

        print(f"\n[{i+1}] {cat}: {q}")

        for tech, builder in TECHNIQUES.items():
            prompt = builder(q, cat) if tech == "Few-Shot" else builder(q)
            response = ask_llm(prompt)

            print(f"  → {tech}: {response[:80]}")

            ws.append([
                i + 1, timestamp, cat, q, a,
                tech, prompt, response
            ])

    wb.save(EXCEL_FILE)

    print("\n Batch complete ")



if __name__ == "__main__":
    main()